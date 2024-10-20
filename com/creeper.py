import os.path
import re
import shutil
import time
from datetime import datetiwme
from random import random, randint
import requests
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.workbook import Workbook
from lxml import etree
from com.example.Main import open_and_slide

class house:
    def __init__(self,house_id, total_price, price, res_quarters, location, area, get_time,url):
        self.house_id = house_id
        self.total_price = total_price
        self.price = price
        self.res_quarters = res_quarters
        self.location = location
        self.area = area
        self.get_time = get_time
        self.url = url

    def show(self):
        print(self.house_id)
        print(self.total_price)
        print(self.price)
        print(self.res_quarters)
        print(self.location)
        print(self.area)
        print(self.get_time)
        print(self.url)

#时间格式美化
def date_amender(*args:str):
    l = list()
    if len(args) != 0:
        for ele in args:
            ele = str(ele)
            if len(ele) < 2:
                l.append("0" + ele)
            else:
                l.append(ele)
    return l

headers = {
    "user-agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36 Edg/129.0.0.0"
}

#--------------xlsx操作-----------------#
side = Side(
    style="medium",
    color="000000",
)

while True:
    try:
        now =  datetime.now()
        month, day = date_amender(now.month,now.day)
        #如果没有文件先创建文件
        if not os.path.exists(r"D:\房源数据.xlsx"):
            wb = Workbook()
            # 保存工作簿到指定路径
            wb.save(r"D:\房源数据.xlsx")
            print("不存在源文件现已创建")
        shutil.copy(r"D:\房源数据.xlsx",f"D:\房源数据{month}-{day}备份.xlsx")
        wb = load_workbook(r"D:\房源数据.xlsx")
        ws = None
        if (month + "-" + day) in wb.sheetnames:
            ws = wb[month + "-" + day]
            print(f"已存在{month + "-" + day}表")

        else:
            ws = wb.create_sheet(month + "-" + day)
            print(f"新创建了{month + "-" + day}表")
            ws["A1"] = "数据获取时间" + month + "-" + day
            ws["A1"].font = Font(bold=True,size=20)
            ws["A1"].alignment = Alignment(horizontal='center', vertical='center')
            ws["A1"].fill = PatternFill(
                patternType="solid",
                fgColor="ffff00",
                bgColor="ffff00",
            )
            ws["A1"].border = Border(left=side, top=side,right=side,bottom=side)
            ws.merge_cells("A1:H2")
            list_data = ["房子id","总价格","单位面积价格","所在小区","地址","面积","爬取时间","网页链接"]
            for i in range(1,9):
                tcell = ws.cell(row=3, column=i, value=list_data[i - 1])
                tcell.font = Font(bold=True)
                tcell.border = Border(left=side, top=side,right=side,bottom=side)
            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["C"].width = 13.5
            ws.column_dimensions["D"].width = 18
            ws.column_dimensions["E"].width = 19
            ws.column_dimensions["G"].width = 14.5
        #--------------xlsx操作----------------/#
        #测试已经存在多少数据
        start_idx = 3
        while ws.cell(row=start_idx,column=1).value is not None:
            start_idx += 1
        #-----爬虫操作---------#
        start_idx = ((start_idx - 4) // 60) + 1
            #爬取时间、网页链接、ID、总价格、单位面积价格、小区、片区、地块、面积、等关键数据

        if (start_idx >= 39):
            print("当日数据量已经完成")
            break
        for i in range(start_idx, 39):
            print(f"当前是第{i}页数据")
            target_url = f"https://wenzhou.anjuke.com/sale/ruianwz/o4-p{i}/?from=HomePage_TopBar/"
            response = requests.get(target_url, headers=headers)
            tree = etree.HTML(response.text)
            res = list(tree.xpath('//*[@id="esfMain"]/section/section[3]/section[1]/section[2]/div/a/@href'))
            now = datetime.now()#当前时间
            print(f"第{i}页一共有{len(res)}个子页面")
            if (len(res) == 0):
                print("未找到数据，可能被检测了")
                raise IndexError
            house_history = list()
            print(res)
            idx = 1
            for link in res:
                print(f"第{i}个页面,第{idx}次访问{link}")
                response2 = requests.get(url=link, headers=headers)
                tree = etree.HTML(response2.text)
                house_id = tree.xpath('//*[@id="__layout"]/div/div[3]/div[1]/div[4]/div[2]/text()')[0].lstrip(
                    "房屋编码： ")  # 房屋id

                total_price = ''.join(tree.xpath("//div[@class='maininfo-price-wrap']/*/text() | //*[@class='maininfo-noprice']/text()"))
                total_price = total_price.strip()
                if (total_price == "暂无售价"):
                    print("该房子已下架")
                    continue

                price = tree.xpath("//div[@class='maininfo-avgprice-price']/text()")[0]
                price = price.strip()

                res_quarters = tree.xpath("//*[text()='所属小区']/following-sibling::a[1]/text()")[0]
                res_quarters = res_quarters.strip()

                location = ' '.join(
                    tree.xpath("//*[text()='所属区域']/following-sibling::span/a/text()"))
                location = location.strip()

                area = ''.join(tree.xpath(
                    "//*[@class='maininfo-model-item maininfo-model-item-2']/*[@class='maininfo-model-strong']/*/text()"))
                area = area.strip()

                now = datetime.now()#获取爬取时间
                get_time = ':'.join(date_amender(now.hour,now.minute,now.second))

                new_house = house(house_id, total_price, price, res_quarters, location, area, get_time,link)
                house_history.append(new_house)
                time.sleep(randint(3,5))
                idx += 1

            sour_data = sorted(house_history,key = lambda e: e.house_id)
            offset = 4 + (i - 1) * 60
            for nowrow in range(offset, offset + len(sour_data)):  # ele是houseid:house
                ele = sour_data[nowrow - offset]
                ws.cell(row=nowrow, column=1, value=ele.house_id)
                ws.cell(row=nowrow, column=2, value=ele.total_price)
                ws.cell(row=nowrow, column=3, value=ele.price)
                ws.cell(row=nowrow, column=4, value=ele.res_quarters)
                ws.cell(row=nowrow, column=5, value=ele.location)
                ws.cell(row=nowrow, column=6, value=ele.area)
                ws.cell(row=nowrow, column=7, value=ele.get_time)
                ws.cell(row=nowrow, column=8, value=ele.url)

            for i in range(offset, len(sour_data) + offset):
                for j in range(1, 9):
                    tcell = ws.cell(row=i, column=j)
                    tcell.border = Border(left=side, top=side, right=side, bottom=side)
            wb.save(r"D:\房源数据.xlsx")
            shutil.copy(r"D:\房源数据.xlsx", f"D:\房源数据{month}-{day}备份.xlsx")
            # time.sleep(randint(10, 15))
        #-----爬虫操作---------#

        # ----美化格式-----#

    except Exception as e:
        print("遇到错误，需要验证码")
        print(e)
        #获取验证码链接
        url = "https://wenzhou.anjuke.com/sale/?from=HomePage_TopBar"
        open_and_slide(url)

