from datetime import datetime

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Side, Alignment, Font, PatternFill, Border

from com.creeper import date_amender

now = datetime.now()
month, day = date_amender(now.month, now.day)
wb = load_workbook(r"D:\ww.xlsx")
ws = wb.create_sheet(month + "-" + day)

side = Side(
    style="medium",
    color="000000",
)
ws["A1"] = "数据获取时间" + month + "-" + day
ws["A1"].font = Font(bold=True, size=20)
ws["A1"].alignment = Alignment(horizontal='center', vertical='center')
ws["A1"].fill = PatternFill(
    patternType="solid",
    fgColor="ffff00",
    bgColor="ffff00",
)
ws["A1"].border = Border(left=side, top=side, right=side, bottom=side)
ws.merge_cells("A1:H2")
list_data = ["房子id", "总价格", "单位面积价格", "所在小区", "地址", "面积", "爬取时间", "网页链接"]
for i in range(1, 9):
    tcell = ws.cell(row=3, column=i, value=list_data[i - 1])
    tcell.font = Font(bold=True)


ws.column_dimensions["A"].width = 18
ws.column_dimensions["C"].width = 13.5
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 19
ws.column_dimensions["G"].width = 14.5

wb.save(r"D:\ww.xlsx")
