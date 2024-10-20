from turtledemo.penrose import start

from openpyxl.reader.excel import load_workbook

wb = load_workbook(r"D:\666.xlsx")
ws = None
if "Sheet3" in wb.sheetnames:
    ws = wb["Sheet3"]

start = 1
while ws.cell(row=start,column=1).value is not None:
    start += 1

print(start)
print(ws["A1"])