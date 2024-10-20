from openpyxl import Workbook



wb = Workbook()#创建工作簿
ws2 = wb.create_sheet("财报")#创建工作表
x = 0
for i in range(1,11):
    for j in range(1,11):
        tcell = ws2.cell(row=i, column=j, value=x)#访问单元格
        # print(tcell.value)#单元格的值
        # print(tcell.coordinate)
        # print(tcell.column)
        # print(tcell.row)
        # print(tcell.column_letter)#字母形式的行
        x += 1
# print(wb.sheetnames)#输出工作簿中所有工作表
print(ws2["A:C"])#返回A-C行的单元格对象列表
print(ws2[1:3])#返回1-3列的单元格对象列表
print(ws2[1])#返回1列的单元格对象列表
print(ws2["A"])#返回A行的单元格对象列表
wb.save(r"D:\text.xlsx")#保存